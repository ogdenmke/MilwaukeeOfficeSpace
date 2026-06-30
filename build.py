#!/usr/bin/env python3
"""
Build script for Ogden Office Space website.
Reads Ogden_Office_Listings.xlsx and produces JSON files the site consumes.
Attempts to geocode any buildings missing lat/long via OpenStreetMap Nominatim.

Usage:
    python3 build.py
"""

import json
import os
import sys
import time
import urllib.request
import urllib.parse
import openpyxl

XLSX_PATH = os.path.join(os.path.dirname(__file__), "Ogden_Office_Listings.xlsx")
DATA_DIR = os.path.join(os.path.dirname(__file__), "site", "data")

NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"

FALLBACK_COORDS = {
    "1661 N Water St": (43.0455, -87.9103),
    "757-765 N Broadway & E Mason St": (43.0408, -87.9085),
    "5555 N Port Washington Rd": (43.1132, -87.9105),
    "205 W Highland Ave & N Old World Third St": (43.0443, -87.9147),
    "3600 W Pierce St": (43.0156, -87.9553),
    "753-757 N Water St": (43.0395, -87.9105),
    "5205 N Ironwood Ln": (43.1070, -87.9186),
    "7810 W Good Hope Rd": (43.1512, -87.9946),
    "1730 W North Ave": (43.0559, -87.9310),
}


def clean_val(v):
    if v is None:
        return None
    # Convert integer-like floats (e.g. floor=1.0 → "1", sf=1141.0 → "1141")
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    s = str(v).strip()
    return s if s else None


def read_sheet(wb, sheet_name):
    ws = wb[sheet_name]
    raw_headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row):
            continue
        d = {}
        for h, v in zip(raw_headers, row):
            if h is None:
                continue
            cv = clean_val(v)
            if cv is not None:
                d[h] = cv
        if d:
            rows.append(d)
    return rows


def geocode(address, city, state, zip_code):
    query = f"{address}, {city}, {state} {zip_code}"
    params = urllib.parse.urlencode({
        "q": query,
        "format": "json",
        "limit": 1,
        "countrycodes": "us",
    })
    url = f"{NOMINATIM_URL}?{params}"
    req = urllib.request.Request(url, headers={"User-Agent": "OgdenOfficeSpaceBuild/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read())
            if data:
                return float(data[0]["lat"]), float(data[0]["lon"])
    except Exception as e:
        print(f"  Geocoding error for '{query}': {e}")
    return None, None


def main():
    if not os.path.exists(XLSX_PATH):
        print(f"ERROR: {XLSX_PATH} not found.")
        sys.exit(1)

    os.makedirs(DATA_DIR, exist_ok=True)
    wb = openpyxl.load_workbook(XLSX_PATH)

    buildings = read_sheet(wb, "Buildings")
    suites = read_sheet(wb, "Suites")
    contacts = read_sheet(wb, "Contacts")

    # Floor plan viewer fields — add to STN if not already in the sheet
    for b in buildings:
        if b.get("building_id") == "STN":
            b.setdefault("floor_count", "5")
            b.setdefault("lobby_floors", "1,2")

    STN_FP = {
        "Suite 306": {"fp_x": "5",  "fp_y": "5",  "fp_w": "90", "fp_h": "90"},
        "Suite 400": {"fp_x": "5",  "fp_y": "5",  "fp_w": "42", "fp_h": "44"},
        "Suite 401": {"fp_x": "53", "fp_y": "5",  "fp_w": "42", "fp_h": "44"},
        "Suite 402": {"fp_x": "5",  "fp_y": "55", "fp_w": "42", "fp_h": "40"},
        "Suite 403": {"fp_x": "53", "fp_y": "55", "fp_w": "42", "fp_h": "40"},
        "Suite 500": {"fp_x": "5",  "fp_y": "5",  "fp_w": "46", "fp_h": "90"},
        "Suite 502": {"fp_x": "56", "fp_y": "5",  "fp_w": "39", "fp_h": "28"},
        "Suite 504": {"fp_x": "56", "fp_y": "36", "fp_w": "39", "fp_h": "28"},
        "Suite 509": {"fp_x": "56", "fp_y": "67", "fp_w": "39", "fp_h": "28"},
    }
    for s in suites:
        if s.get("building_id") == "STN":
            fp = STN_FP.get(s.get("suite_number", ""))
            if fp:
                for k, v in fp.items():
                    s.setdefault(k, v)

    unresolved = []
    for b in buildings:
        if not b.get("latitude") or not b.get("longitude"):
            print(f"Geocoding: {b['building_name']} — {b['address']}, {b['city']}, {b['state']} {b['zip']}")
            lat, lon = geocode(b["address"], b["city"], b["state"], b["zip"])
            if lat and lon:
                b["latitude"] = str(lat)
                b["longitude"] = str(lon)
                print(f"  → {lat}, {lon}")
                time.sleep(1.1)  # Nominatim rate limit: 1 req/sec
            elif b["address"] in FALLBACK_COORDS:
                lat, lon = FALLBACK_COORDS[b["address"]]
                b["latitude"] = str(lat)
                b["longitude"] = str(lon)
                print(f"  → {lat}, {lon} (fallback)")
            else:
                unresolved.append(f"  {b['building_name']}: {b['address']}, {b['city']}, {b['state']} {b['zip']}")
                print(f"  → COULD NOT RESOLVE — fill in lat/long manually")

    with open(os.path.join(DATA_DIR, "buildings.json"), "w") as f:
        json.dump(buildings, f, indent=2)
    with open(os.path.join(DATA_DIR, "suites.json"), "w") as f:
        json.dump(suites, f, indent=2)
    with open(os.path.join(DATA_DIR, "contacts.json"), "w") as f:
        json.dump(contacts, f, indent=2)

    print(f"\nDone. Wrote {len(buildings)} buildings, {len(suites)} suites, {len(contacts)} contacts to {DATA_DIR}/")

    if unresolved:
        print("\n⚠  Could not geocode the following — add lat/long manually in the spreadsheet:")
        for u in unresolved:
            print(u)


if __name__ == "__main__":
    main()
